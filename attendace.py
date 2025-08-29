import io
import re
from datetime import time
from typing import Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# --- Employee mapping ---
employees = {
    1: "Ishmal", 2: "Owais", 3: "Neha", 4: "Sarah", 6: "Musfira", 7: "Ayesha",
    8: "Junaid", 9: "Abd-ur-Rehman", 10: "Samra", 13: "Shahida", 15: "Seerat",
    16: "Usman", 17: "Mahad", 18: "Eman", 19: "Izaz", 20: "Kiran", 21: "Hammad",
    24: "Faiza", 26: "Laiba", 29: "Ushna", 30: "Ali", 31: "Adnan", 32: "Yaseen",
    33: "Abbas"
}

CHECKIN_THRESHOLD = time(9, 2, 0)       # 9:02 AM
CHECKOUT_THRESHOLD = time(17, 0, 0)     # 5:00 PM

# --- Streamlit setup ---
st.set_page_config(page_title="Attendance Dashboard", page_icon="📊", layout="wide")
st.title("📊 Attendance Processor")
st.markdown("Upload your ZKTeco `.dat` / `.txt` / `.csv` export to view attendance summary.")

# --- Sidebar controls ---
st.sidebar.header("⚙️ Settings")
weekends_off = st.sidebar.checkbox("Ignore weekends (Sat/Sun)", value=False)
uploaded = st.sidebar.file_uploader("Upload .dat / .txt / .csv file", type=["dat","txt","csv","log"])

# --- Helpers ---
CANDIDATE_USER_COLS = ["userid", "user_id", "pin", "enrollid", "empid", "id"]
CANDIDATE_TIME_COLS = ["timestamp", "time", "datetime", "logtime", "punch time", "punch_time"]

def _try_read_csv(buf: io.BytesIO) -> Optional[pd.DataFrame]:
    buf.seek(0)
    for sep in ["\t", ",", ";", "|"]:
        buf.seek(0)
        try:
            df = pd.read_csv(buf, sep=sep, engine="python")
            if df.shape[1] >= 2:
                return df
        except Exception:
            pass
    buf.seek(0)
    try:
        return pd.read_csv(buf, delim_whitespace=True, engine="python", header=None)
    except Exception:
        return None

def _coerce_encoding(file_bytes: bytes) -> io.BytesIO:
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1", errors="ignore")
    return io.BytesIO(text.encode("utf-8"))

def _detect_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    lower_cols = {c.lower(): c for c in df.columns.astype(str)}
    user_col, ts_col = None, None
    for k in CANDIDATE_USER_COLS:
        if k in lower_cols:
            user_col = lower_cols[k]; break
    for k in CANDIDATE_TIME_COLS:
        if k in lower_cols:
            ts_col = lower_cols[k]; break
    if user_col is None and ts_col is None and df.shape[1] >= 2:
        user_col, ts_col = df.columns[0], df.columns[1]
    return user_col, ts_col

def _extract_dataframe(file_bytes: bytes) -> pd.DataFrame:
    buf = _coerce_encoding(file_bytes)
    raw = _try_read_csv(buf)
    if raw is None or raw.empty:
        st.error("Could not parse the file. Check format or try another export.")
        st.stop()
    if any(not isinstance(c, str) for c in raw.columns):
        raw.columns = [f"col_{i+1}" for i in range(raw.shape[1])]
    user_col, ts_col = _detect_columns(raw)
    if user_col is None or ts_col is None:
        st.error(f"Could not detect UserID/Timestamp columns. Found: {list(raw.columns)}")
        st.stop()
    df = raw[[user_col, ts_col]].copy()
    df.columns = ["UserID", "Timestamp"]
    df["UserID"] = df["UserID"].astype(str).str.strip()

    def parse_dt(x):
        x = str(x).strip()
        x = re.sub(r"\s+", " ", x)
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%m/%d/%Y %H:%M:%S",
                    "%Y/%m/%d %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M",
                    "%d/%m/%Y %H:%M", "%m/%d/%Y %H:%M"):
            try:
                return pd.to_datetime(x, format=fmt)
            except Exception:
                pass
        return pd.to_datetime(x, errors="coerce")

    df["Timestamp"] = df["Timestamp"].apply(parse_dt)
    df = df.dropna(subset=["Timestamp"]).copy()
    df["Date"] = df["Timestamp"].dt.date
    if weekends_off:
        df = df[df["Timestamp"].dt.weekday < 5]
    return df

def _summarize_attendance(df: pd.DataFrame, late_t: time, early_checkout_t: time) -> pd.DataFrame:
    first = df.groupby(["UserID", "Date"], as_index=False)["Timestamp"].min().rename(columns={"Timestamp": "CheckIn"})
    last = df.groupby(["UserID", "Date"], as_index=False)["Timestamp"].max().rename(columns={"Timestamp": "CheckOut"})
    merged = pd.merge(first, last, on=["UserID", "Date"], how="inner")
    merged["UserID_int"] = pd.to_numeric(merged["UserID"], errors="coerce")
    merged["Name"] = merged["UserID_int"].map(employees)
    merged = merged.dropna(subset=["Name"]).copy()
    merged["CheckIn"] = merged["CheckIn"].dt.time
    merged["CheckOut"] = merged["CheckOut"].dt.time
    merged["Status"] = merged["CheckIn"].apply(lambda t: "Late" if t > late_t else "On Time")
    merged["Minutes Late"] = [
        max(0, (t.hour - late_t.hour) * 60 + (t.minute - late_t.minute))
        for t in merged["CheckIn"]
    ]
    merged["Early Checkout"] = merged["CheckOut"].apply(lambda t: "Yes" if t < early_checkout_t else "No")
    merged["Minutes Early"] = [
        max(0, (early_checkout_t.hour - t.hour) * 60 + (early_checkout_t.minute - t.minute))
        for t in merged["CheckOut"]
    ]
    merged = merged.sort_values(by=["UserID_int", "Date"])
    merged.drop(columns=["UserID_int"], inplace=True)
    return merged[["UserID", "Name", "Date", "CheckIn", "CheckOut", "Status", "Minutes Late", "Early Checkout", "Minutes Early"]]

def _to_styled_excel(df: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    df.to_excel(bio, index=False, sheet_name="Attendance")
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb.active
    header = [c.value for c in ws[1]]

    # --- Conditional formatting ---
    red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    blue = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")

    last_row = ws.max_row
    last_col_letter = ws.cell(row=1, column=ws.max_column).column_letter

    # Late
    if "Status" in header:
        idx = header.index("Status") + 1
        col_letter = ws.cell(row=1, column=idx).column_letter
        rule = FormulaRule(formula=[f'${col_letter}2="Late"'], fill=red)
        ws.conditional_formatting.add(f"A2:{last_col_letter}{last_row}", rule)
    # Early Checkout
    if "Early Checkout" in header:
        idx = header.index("Early Checkout") + 1
        col_letter = ws.cell(row=1, column=idx).column_letter
        rule = FormulaRule(formula=[f'${col_letter}2="Yes"'], fill=yellow)
        ws.conditional_formatting.add(f"A2:{last_col_letter}{last_row}", rule)

    # Adjust column widths
    for col in ws.columns:
        max_len, col_letter = 0, col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 40)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

# --- Main Flow ---
if uploaded:
    file_bytes = uploaded.read()
    df_raw = _extract_dataframe(file_bytes)
    st.success(f"Parsed {len(df_raw):,} log rows successfully!")

    summary = _summarize_attendance(df_raw, CHECKIN_THRESHOLD, CHECKOUT_THRESHOLD)

    # --- Metrics ---
    total_logs = len(summary)
    total_late = (summary["Status"] == "Late").sum()
    total_on_time = (summary["Status"] == "On Time").sum()
    total_early = (summary["Early Checkout"] == "Yes").sum()

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("📄 Total Records", f"{total_logs}")
    col2.metric("⏰ Total Late", f"{total_late}")
    col3.metric("✅ Total On Time", f"{total_on_time}")
    col4.metric("⚠️ Early Checkout", f"{total_early}")

    st.markdown("---")

    # --- Detailed Attendance ---
    with st.expander("🗂️ View Detailed Attendance Table"):
        st.dataframe(summary)

    # --- Download Excel ---
    excel_bytes = _to_styled_excel(summary)
    st.download_button(
        "⬇️ Download Excel",
        data=excel_bytes,
        file_name="attendance_marked.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")

    # --- Summary Counts ---
    late_counts = summary.groupby("Name")["Status"].apply(lambda x: (x=="Late").sum()).reset_index().rename(columns={"Status":"Total Lates"})
    early_counts = summary.groupby("Name")["Early Checkout"].apply(lambda x: (x=="Yes").sum()).reset_index().rename(columns={"Early Checkout":"Total Early Checkout"})

    # Leaves
    dates_in_data = summary["Date"].unique()
    all_data = pd.MultiIndex.from_product([employees.values(), dates_in_data], names=["Name","Date"])
    full_df = pd.DataFrame(index=all_data).reset_index()
    merged_summary = pd.merge(full_df, summary[["Name","Date"]], on=["Name","Date"], how="left", indicator=True)
    leave_counts = merged_summary.groupby("Name")["_merge"].apply(lambda x: (x=="left_only").sum()).reset_index().rename(columns={"_merge":"Total Leaves"})

    # Merge all counts
    summary_counts = pd.DataFrame({"Name": list(employees.values())})
    summary_counts = summary_counts.merge(late_counts, on="Name", how="left").fillna(0)
    summary_counts = summary_counts.merge(early_counts, on="Name", how="left").fillna(0)
    summary_counts = summary_counts.merge(leave_counts, on="Name", how="left").fillna(0)
    summary_counts[["Total Lates","Total Early Checkout","Total Leaves"]] = summary_counts[["Total Lates","Total Early Checkout","Total Leaves"]].astype(int)

    # --- Display Summary Table with Correct Highlighting ---
    with st.expander("📋 Total Late, Early Checkout & Leave Count of Employees"):
        def highlight_counts(row):
            return [
                "",  # Name
                'background-color: #FFCDD2' if row["Total Lates"] > 0 else '',
                'background-color: #FFF59D' if row["Total Early Checkout"] > 0 else '',
                'background-color: #BBDEFB' if row["Total Leaves"] > 0 else ''
            ]
        st.dataframe(summary_counts.style.apply(highlight_counts, axis=1))
